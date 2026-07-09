import os
import re
import sys
import time
import json
import random
import urllib.request
import urllib.error
import unicodedata
from datetime import datetime, timedelta
import openpyxl
import builtins

# Sobrescreve print para sempre usar flush=True
def print(*args, **kwargs):
    kwargs['flush'] = True
    builtins.print(*args, **kwargs)

# Configurações
EXCEL_PATH = 'Dicionario finanças - NEW.xlsx'
POSTS_DIR = 'content/posts'
GROQ_API_KEY = os.environ.get("GROQ_API_KEY", "")
MODEL_NAME = 'llama-3.3-70b-versatile'
MAX_RETRIES = 8
INITIAL_BACKOFF = 10  # segundos para backoff de RPM

# Listas de meses em português para o frontmatter
MONTHS_PT = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']

def to_slug(text):
    if not text:
        return ""
    # Remover acentos
    text = ''.join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')
    text = text.lower()
    # Substituir caracteres não alfanuméricos por hífen
    text = re.sub(r'[^a-z0-9\s-]', '', text)
    # Substituir múltiplos espaços/hífens por um único hífen
    text = re.sub(r'\s+', '-', text)
    text = re.sub(r'-+', '-', text)
    return text.strip('-')[:60]

def get_existing_slugs():
    if not os.path.exists(POSTS_DIR):
        os.makedirs(POSTS_DIR)
        return set()
    files = os.listdir(POSTS_DIR)
    slugs = set()
    for f in files:
        if f.endswith('.md'):
            slugs.add(f[:-3])
    return slugs

def call_groq_with_retry(prompt):
    url = "https://api.groq.com/openai/v1/chat/completions"
    headers = {
        "Authorization": f"Bearer {GROQ_API_KEY}",
        "Content-Type": "application/json",
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    data = {
        "model": MODEL_NAME,
        "messages": [
            {
                "role": "system",
                "content": "Você é um especialista altamente experiente em finanças corporativas, contabilidade e modelagem financeira. Responda estritamente seguindo o formato de delimitação solicitado."
            },
            {
                "role": "user",
                "content": prompt
            }
        ],
        "temperature": 0.7,
        "max_tokens": 3000
    }
    
    req_body = json.dumps(data).encode('utf-8')
    
    attempt = 0
    while True:
        try:
            req = urllib.request.Request(url, data=req_body, headers=headers, method='POST')
            with urllib.request.urlopen(req, timeout=60) as response:
                res_data = json.loads(response.read().decode('utf-8'))
                return res_data['choices'][0]['message']['content']
        except urllib.error.HTTPError as e:
            body = e.read().decode('utf-8') if e.fp else ""
            print(f"⚠️ Erro HTTP {e.code} na tentativa {attempt+1}: {e.reason}")
            
            if e.code == 429:
                # Verificar se é cota diária (TPD) ou limite por minuto (RPM/TPM)
                is_daily_limit = 'tokens per day' in body.lower() or 'TPD' in body
                
                if is_daily_limit:
                    # Calcular segundos até meia-noite UTC + 90s de folga
                    now_utc = datetime.utcnow()
                    next_midnight = datetime(now_utc.year, now_utc.month, now_utc.day) + timedelta(days=1)
                    wait_seconds = (next_midnight - now_utc).total_seconds() + 90
                    wait_minutes = wait_seconds / 60
                    print(f"⏳ COTA DIÁRIA (TPD) ESGOTADA. Aguardando {wait_minutes:.1f} minutos até reset da meia-noite UTC...")
                    print(f"   Reiniciando geração em: {next_midnight.strftime('%Y-%m-%d %H:%M')} UTC")
                    time.sleep(wait_seconds)
                    attempt = 0  # Resetar contador de tentativas após espera do TPD
                else:
                    # Rate limit de curto prazo (RPM/TPM) — backoff exponencial
                    if attempt >= MAX_RETRIES:
                        raise Exception(f"Falha após {MAX_RETRIES} tentativas por RPM rate limit.")
                    sleep_time = INITIAL_BACKOFF * (2 ** min(attempt, 5)) + random.uniform(1, 5)
                    print(f"⏳ Rate limit por minuto. Aguardando {sleep_time:.1f}s...")
                    time.sleep(sleep_time)
                    attempt += 1
            elif e.code >= 500:
                sleep_time = 10 * (attempt + 1)
                print(f"⏳ Erro no servidor Groq. Aguardando {sleep_time}s...")
                time.sleep(sleep_time)
                attempt += 1
                if attempt >= MAX_RETRIES:
                    raise e
            else:
                raise e
        except Exception as e:
            if '429' in str(e) or 'rate limit' in str(e).lower():
                raise e  # Já tratado acima
            print(f"⚠️ Erro inesperado: {e}")
            attempt += 1
            if attempt >= MAX_RETRIES:
                raise e
            time.sleep(10)

def parse_groq_content(raw_content):
    sections = {}
    # Split using the pattern ===SECTION===
    parts = re.split(r'\n?===([A-Z]+)===\n?', raw_content)
    # parts will be like ['', 'TITLE', 'title content', 'EXCERPT', 'excerpt content', ...]
    for i in range(1, len(parts), 2):
        if i + 1 < len(parts):
            sections[parts[i]] = parts[i+1].strip()
    return sections

def load_excel_terms():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    terms_list = []
    
    for sheet_name in wb.sheetnames:
        if sheet_name in ['INDEX', 'Changelog', 'PLANILHA COMPLETA']:
            continue
        
        ws = wb[sheet_name]
        print(f"Carregando folha: {sheet_name}")
        
        # Ler cabeçalho
        header = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        
        # Mapear posições de coluna
        # Colunas esperadas: Term (EN), Termo (PT-BR), Fórmula, Explicação, Exemplo, Categoria, Nível, Termos Relacionados, Armadilhas / Red Flags
        col_indices = {}
        for idx, col in enumerate(header):
            if not col:
                continue
            col_clean = col.lower().strip()
            if 'term (en)' in col_clean or 'en-us' in col_clean or 'english' in col_clean:
                col_indices['term_en'] = idx
            elif 'termo (pt-br)' in col_clean or 'pt-br' in col_clean or 'portugues' in col_clean:
                col_indices['term_pt'] = idx
            elif 'fórmula' in col_clean or 'formula' in col_clean:
                col_indices['formula'] = idx
            elif 'explicação' in col_clean or 'explicacao' in col_clean or 'explanation' in col_clean:
                col_indices['explicacao'] = idx
            elif 'exemplo' in col_clean or 'example' in col_clean:
                col_indices['exemplo'] = idx
            elif 'categoria' in col_clean or 'category' in col_clean:
                col_indices['categoria'] = idx
            elif 'nível' in col_clean or 'nivel' in col_clean or 'level' in col_clean:
                col_indices['nivel'] = idx
            elif 'relacionados' in col_clean or 'related' in col_clean:
                col_indices['termos_relacionados'] = idx
            elif 'armadilhas' in col_clean or 'red flags' in col_clean or 'danger' in col_clean:
                col_indices['armadilhas'] = idx
                
        # Verificar mapeamento básico
        if 'term_pt' not in col_indices or 'term_en' not in col_indices:
            print(f"⚠️ Ignorando folha '{sheet_name}' devido a colunas de termo não identificadas: {header}")
            continue
            
        for row in list(ws.iter_rows(min_row=2)):
            # Validar se o primeiro campo (#) é um número para ignorar linhas de rodapé ou vazias
            hash_val = row[0].value
            if hash_val is None:
                continue
            try:
                float(hash_val)
            except ValueError:
                continue # não é um número, pula
                
            term_pt = row[col_indices['term_pt']].value
            term_en = row[col_indices['term_en']].value
            
            if not term_pt:
                continue
                
            term_data = {
                'sheet': sheet_name,
                'term_pt': term_pt,
                'term_en': term_en or "",
                'formula': row[col_indices.get('formula')].value if 'formula' in col_indices else "",
                'explicacao': row[col_indices.get('explicacao')].value if 'explicacao' in col_indices else "",
                'exemplo': row[col_indices.get('exemplo')].value if 'exemplo' in col_indices else "",
                'categoria': row[col_indices.get('categoria')].value if 'categoria' in col_indices else sheet_name,
                'nivel': row[col_indices.get('nivel')].value if 'nivel' in col_indices else "Intermediário",
                'termos_relacionados': row[col_indices.get('termos_relacionados')].value if 'termos_relacionados' in col_indices else "",
                'armadilhas': row[col_indices.get('armadilhas')].value if 'armadilhas' in col_indices else ""
            }
            
            # Garantir que todos os campos vazios sejam strings limpas
            for k in term_data:
                if term_data[k] is None:
                    term_data[k] = ""
            
            terms_list.append(term_data)
            
    print(f"📊 Total de termos carregados da planilha: {len(terms_list)}")
    return terms_list

def main():
    print("🚀 Iniciando Gerador de Posts para o Blog...")
    
    # Adicionar suporte a limite de posts por argumento
    limit = None
    if len(sys.argv) > 1:
        for arg in sys.argv[1:]:
            if arg.startswith('--limit='):
                try:
                    limit = int(arg.split('=')[1])
                    print(f"Limitando a geração para no máximo {limit} posts.")
                except ValueError:
                    pass
            elif arg == '-h' or arg == '--help':
                print("Uso: python3 scripts/generate-posts.py [--limit=N]")
                return
    
    # 1. Carregar slugs de posts existentes
    existing_slugs = get_existing_slugs()
    print(f"📂 Posts existentes na pasta: {len(existing_slugs)}")
    
    # 2. Carregar termos da planilha
    all_terms = load_excel_terms()
    
    # 3. Filtrar termos que já possuem post
    pending_terms = []
    for term in all_terms:
        # Tentar cruzar pelo slug do termo em PT ou EN
        slug_pt = to_slug(term['term_pt'])
        slug_en = to_slug(term['term_en'])
        
        # Também verificar se o slug tem variações comuns (como prefixos 'o-que-e', 'como-calcular', etc.)
        # Os arquivos costumam ser salvos com prefixos. Vamos verificar se algum slug existente CONTÉM o slug do termo
        already_exists = False
        if slug_pt in existing_slugs or slug_en in existing_slugs:
            already_exists = True
        else:
            # Varre os slugs existentes buscando correspondência
            for existing in existing_slugs:
                if slug_pt in existing or (slug_en and slug_en in existing):
                    already_exists = True
                    break
                    
        if not already_exists:
            pending_terms.append(term)
            
    print(f"🔍 Termos pendentes para geração: {len(pending_terms)}")
    
    if not pending_terms:
        print("✅ Todos os termos já possuem posts gerados! Finalizando.")
        return
        
    # Limitar para gerar no máximo os termos pendentes (ou os 300 solicitados pelo usuário)
    # Como o usuário pediu 300 posts no total, e já temos 62, a geração de 240-278 cobrirá a meta.
    # Vamos rodar sequencialmente.
    
    # Embaralhar para não gerar apenas uma categoria por vez, dando variedade nas datas
    random.shuffle(pending_terms)
    
    if limit is not None:
        pending_terms = pending_terms[:limit]
        print(f"Subconjunto reduzido para geração: {len(pending_terms)} posts.")
    
    count_generated = 0
    start_date = datetime.now() - timedelta(days=1) # começar de ontem e ir retrocedendo para dar histórico ao blog
    
    for i, term in enumerate(pending_terms):
        print(f"\n--- 📝 [{i+1}/{len(pending_terms)}] Gerando: {term['term_pt']} ({term['term_en']}) ---")
        
        prompt = f"""Você é um especialista altamente experiente em finanças corporativas, contabilidade e modelagem financeira.
Escreva um post de blog técnico, extremamente detalhado, didático e aprofundado em Português do Brasil sobre o termo financeiro: "{term['term_pt']}" (também conhecido como "{term['term_en']}").

Aqui estão os dados técnicos oficiais do termo para você incorporar no post de forma profissional:
- Categoria: {term['categoria']}
- Nível de complexidade: {term['nivel']}
- Fórmula: {term['formula']}
- Explicação básica: {term['explicacao']}
- Exemplo prático: {term['exemplo']}
- Armadilhas e Sinais de Alerta (Red Flags) comuns: {term['armadilhas']}
- Termos relacionados: {term['termos_relacionados']}

O post deve ser estruturado de forma rica para profissionais, gestores e estudantes de finanças. Use uma linguagem clara, mas com alto nível de precisão técnica. Vá direto ao ponto, evitando introduções clichês ou genéricas sobre o mercado financeiro.

Sua resposta deve seguir EXATAMENTE o formato abaixo, usando estritamente os separadores indicados para cada seção (===TITLE===, ===EXCERPT===, ===READTIME===, ===CONTENT===):

===TITLE===
[Insira um título de blog profissional e atraente, ex: "O que é {term['term_pt']} e como calcular na prática"]
===EXCERPT===
[Resumo executivo em 1-2 frases explicando a importância e uso deste conceito]
===READTIME===
[Tempo estimado de leitura, ex: 8 min]
===CONTENT===
[Escreva o conteúdo do post em formato Markdown completo.
Instruções de estrutura obrigatórias:
1. Comece com um subtítulo ## O que é o {term['term_pt']}? e explique o conceito de forma rica, seu propósito e quem o utiliza.
2. Crie uma seção ## A Fórmula e Componentes do {term['term_pt']}. Explique a matemática por trás (se aplicável), desmembrando cada componente de '{term['formula']}'. Se o termo não possuir uma fórmula matemática, explique o framework ou metodologia de cálculo.
3. Adicione uma seção ## Exemplo Prático de Aplicação. Crie um cenário detalhado de uma empresa fictícia onde o leitor acompanha o cálculo passo a passo, a interpretação do resultado e a tomada de decisão gerencial baseada nele. Use tabelas em markdown para ilustrar os números.
4. Adicione uma seção ## Armadilhas e Sinais de Alerta (Red Flags). Apresente detalhadamente as armadilhas comuns citadas: '{term['armadilhas']}' e como evitá-las no dia a dia.
5. Adicione uma seção ## Termos Relacionados e Conclusão. Faça pontes e explicações de como ele interage com '{term['termos_relacionados']}' e encerre com um sumário acionável.

Garanta profundidade máxima e clareza. O conteúdo do Markdown (seção CONTENT) deve ter pelo menos 600 palavras e conter exemplos práticos reais.]
"""
        
        try:
            raw_res = call_groq_with_retry(prompt)
            sections = parse_groq_content(raw_res)
            
            title = sections.get('TITLE', '').strip().replace('"', '\\"')
            excerpt = sections.get('EXCERPT', '').strip().replace('"', '\\"')
            read_time = sections.get('READTIME', '8 min').strip()
            content = sections.get('CONTENT', '').strip()
            
            if not title or not content:
                print("❌ Falha ao parsear resposta da IA. TITLE ou CONTENT ausentes.")
                print(f"Resposta bruta da API:\n{raw_res[:500]}...")
                continue
                
            slug = to_slug(title)
            if not slug:
                slug = to_slug(term['term_pt'])
                
            # Formatar a data para o frontmatter de forma retroativa (um dia a menos por post)
            post_date = start_date - timedelta(days=count_generated)
            date_str = f"{str(post_date.day).zfill(2)} {MONTHS_PT[post_date.month - 1]} {post_date.year}"
            
            # Construir o markdown final
            markdown = f"""---
title: "{title}"
date: "{date_str}"
readTime: "{read_time}"
author: "Tiago Porto"
image: "https://images.unsplash.com/photo-1554224155-6726b3ff858f?auto=format&fit=crop&q=80&w=1200"
excerpt: "{excerpt}"
---

{content}
"""
            
            file_path = os.path.join(POSTS_DIR, f"{slug}.md")
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(markdown)
                
            print(f"✅ Post salvo com sucesso: {file_path}")
            print(f"   Título: {title}")
            print(f"   Data: {date_str} | Slug: {slug}")
            
            count_generated += 1
            
            # Pequeno delay para acalmar a API
            time.sleep(12)
            
        except Exception as e:
            print(f"❌ Erro ao processar o termo '{term['term_pt']}': {e}")
            continue
            
    print(f"\n🎉 Geração concluída! {count_generated} novos posts gerados.")

if __name__ == "__main__":
    main()
